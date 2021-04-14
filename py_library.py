
import openpyxl
#import random
from openpyxl import Workbook
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart import PieChart3D, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.formatting.rule import ColorScaleRule
#from openpyxl.styles import Color
from openpyxl.styles import PatternFill
wb = Workbook()
sheet = wb.active
# Appending Data to Sheet
rows = [
    ["Name", "Sub1", "Sub2", "Sub3", "Total"],
    ["Aniket", 98, 85, 89],
    ["Max", 84, 75, 98],
    ["John", 34, 85, 45],
    ["Rocky", 68, 75, 89],

]

for row in rows:
    sheet.append(row)

    # This command will set filter in sheet
sheet.auto_filter.ref = "A1:F11"
# Formulas for IF/OR condition
sheet["F2"] = '=IF(OR(B2<35,C2<35,D2<35),"Fail","Pass")'
sheet["F3"] = '=IF(OR(B3<35,C3<35,D3<35),"Fail","Pass")'
sheet["F4"] = '=IF(OR(B4<35,C4<35,D4<35),"Fail","Pass")'
sheet["F5"] = '=IF(OR(B5<35,C5<35,D5<35),"Fail","Pass")'
# Vlookup formula to check By name Pass or Fail
sheet["G1"] = "Vlookup Result"
sheet["G2"] = "= VLOOKUP(A2:A5,A1:F5,6,0)"
sheet["G3"] = "= VLOOKUP(A2:A5,A1:F5,6,0)"
sheet["G4"] = "= VLOOKUP(A2:A5,A1:F5,6,0)"
sheet["G5"] = "= VLOOKUP(A2:A5,A1:F5,6,0)"
# Fill color
sheet['A2'].fill = PatternFill(bgColor="ff7f50", fill_type="lightTrellis")
# This will rename
sheetname = "Aniket"
sheet.title = sheetname
'''
# Add sheet to workbook
sheetname = "Day2 Result"
wb.create_sheet(index=1, title=sheetname)
'''
# Count Formula
sheet["H1"] = "Count Formulas"
# COUNT = counts Numeric data only
sheet["H2"] = "=COUNT(A1:G5)"
# CCOUNTA = counts All sheet Cell Without Blank
sheet["H3"] = "=COUNTA(A1:G5)"
# COUNTBLANK = counts only blank cells
sheet["H4"] = "=COUNTBLANK(A1:G5)"
#
#
#
'''
RED = "C70E0F"
Color_scale_rule = ColorScaleRule(start_type="min",
                                  start_color=Color.RED,
                                  end_type="max",
                                  end_color=Color.GREEN)
data_bar_rule = DataBarRule(
    start_type="num", start_value=35, end_type="num", end_value=99, color=Color.green)
sheet.conditional_formatting.add("B2:D5", data_bar_rule)
'''
wb.save(filename="demon.xlsx")
# To Calculate Total


def val(x, y):
    return sheet.cell(row=x, column=y).value


for m in range(2, 6):
    sheet.cell(row=m, column=5).value = val(m, 2) + val(m, 3) + val(m, 4)

wb.save(filename="demon.xlsx")


'''
    # To print Random Values by andom libraries
for row in sheet.iter_rows(min_row=8,
                           max_row=13,
                           min_col=2,
                           max_col=13):
    for cell in row:
        cell.value = random.randrange(5, 100)

    '''
# Adding Barchart to sheet on record in sheet
chart = BarChart3D()
data = Reference(worksheet=sheet,
                 min_row=1,
                 max_row=8,
                 min_col=2,
                 max_col=4)

chart.add_data(data, titles_from_data=True)
chart.title = " MARKS "
sheet.add_chart(chart, "J2")
wb.save(filename="demon.xlsx")

# Adding PIEchart


chart = PieChart3D()
data = Reference(worksheet=sheet,
                 min_row=1,
                 max_row=4,
                 min_col=2,
                 max_col=4)

chart.add_data(data, titles_from_data=True)
chart.title = " MARKS "
# chart.set_categories(data)
sheet.add_chart(chart, "A10")
wb.save(filename="demon.xlsx")


# Adding data cell by cell
sheet["A1"] = "Name"
sheet["A2"] = "Aniket"
sheet["F1"] = "Result"

wb.save(filename="demon.xlsx")


# To read a Workbook
wb = openpyxl.load_workbook("demon.xlsx")
cell = sheet.cell(row=1, column=1)
print(cell)
print("Printing Cell", cell.value)

# To print columns
max_column = sheet.max_column
for i in range(1, max_column+1):
    cell = sheet.cell(row=1, column=i)
    print("Printing columns", cell.value)


# To print Rows
max_row = sheet.max_row
for j in range(1, max_row+1):
    cell = sheet.cell(row=j, column=1)
    print("Printing Rows", cell.value)
