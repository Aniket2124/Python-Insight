import openpyxl

path = "C:\\Users\\VIRAJ\\Desktop\\Aniket\\PythonLearning\\Python-Insight\\Data.xlsx"
# This will load workbook from given path
wb_obj = openpyxl.load_workbook(path)
# wb_obj.sheetnames # this will dislpay <Worksheet "Sheet1">,Sheet1
sheet = wb_obj.active

# Toappend Data In existing Sheet
sheet["d3"] = "Aniket"
sheet["d4"] = "Viraj"
sheet["E1"] = "Emp_Id"
sheet["D1"] = "Emp_Name"
sheet["F1"] = "Target"
sheet["D2"] = "Abhi"
sheet["E2"] = "0123"
sheet["F2"] = "5000"
sheet["F3"] = "6000"
sheet["E3"] = "0124"
sheet["D3"] = "Avi"
wb_obj.save(filename="Data_append.xlsx")
sheet.auto_filter.ref = "A1:F11"  # This command will set filter in sheet
wb_obj.save(filename="Data_append.xlsx")
print(sheet.dimensions)
sheet.title = "Data"  # This will rename
print("sheet name is: " + sheet.title)


for value in sheet.iter_rows(min_row=1, max_col=6, values_only=True):

    print(value)
